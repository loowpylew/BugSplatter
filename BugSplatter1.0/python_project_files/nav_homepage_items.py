import f_driver_init as fdi
import globals as gs
import menu_items as mi
import openpyxl
from shutil import copyfile
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import option_menus as om
import time
import os

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException


### nav_homepage_items ###

# Function to adjust column widths based on the longest string within its corresponding cells
def adjust_column_widths(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


# Function used to list files in directory path passed to it ending at the root folder to which the files are presnet i.e. /source_system, /target_system
def list_files_in_directory(directory_path):
    try:
        # Get the list of all files and directories in the specified directory
        files = []
        for file_name in os.listdir(directory_path):
            # Construct full file path
            full_path = os.path.join(directory_path, file_name)
            # Check if it is a file
            if os.path.isfile(full_path):
                files.append(full_path)
        return files
    except Exception as e:
        print("An error occurred:", e)
        return []


# Function to read User Roles from config file and store in array
def read_user_roles(config_file, sheet_name='User_Roles'):
    workbook = openpyxl.load_workbook(config_file)
    sheet = workbook[sheet_name]
    roles = []
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        roles.append(row[0])
    return roles


# Function to create source files for each user role
def create_source_files(user_roles):

    ## Step 2: Copy the existing Excel file
    original_file = '..\\excel_sheets\\nav_homepage_items\\setup\\homepage_items_setup.xlsx'  # Update with the path to your original Excel file

    i = 1
    copied_files = []
    # We peform this for loop outside of the concurrent pool so that we have all excel sheets to hand.
    for user_role in user_roles:
        print(user_role)
        copied_file = f'..\\excel_sheets\\nav_homepage_items\\source_system\\homepage_items_userrole_{i}.xlsx'  # Update with the path where you want to save the copied file
        
        copied_files.append(copied_file)

        i += 1

        copyfile(original_file, copied_file)

        print("Excel file created: ", copied_file)

    return copied_files

# Function to create target files for each user role
def create_target_files(user_roles): 
    ## Step 2: Copy the existing Excel file
    original_file = '..\\excel_sheets\\nav_homepage_items\\setup\\homepage_items_setup.xlsx'  # Update with the path to your original Excel file

    i = 1
    copied_files = []
    # We peform this for loop outside of the concurrent pool so that we have all excel sheets to hand.
    for user_role in user_roles:
        print(user_role)
        copied_file = f'..\\excel_sheets\\nav_homepage_items\\target_system\\homepage_items_userrole_{i}.xlsx'  # Update with the path where you want to save the copied file
        
        copied_files.append(copied_file)

        i += 1

        copyfile(original_file, copied_file)

        print("Excel file created: ", copied_file)

    return copied_files


# Function to scrape main menu menu items once loaded onto the webpage
def scrape_main_menu_items(target_sys, headless_mode, username, password, driver_response, user_type): 

    gs.loginAndLoadHomePage(target_sys, headless_mode, username, password, driver_response, True)
    gs.checkUserRoleAndSwitch(target_sys, headless_mode, user_type, username, password, driver_response, True)
    
    time.sleep(20)

    # Get the page source and parse it with BeautifulSoup
    page_source = driver_response.page_source
    #print(page_source)
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find all elements with the class 'link_name'
    link_elements = soup.find_all(class_='link_name')
    
    # Extract the text content of each element
    menu_names_set = {link.get_text(strip=True) for link in link_elements}

    # Convert the set to a list
    menu_names_array = list(menu_names_set)

    driver_response.quit()
    
    # Print the list of menu item names
    # print(menu_names_array)

    return menu_names_array


# Function to call the 'scrape_main_menu_menu_items' function on seperate threads based on the usertype read from the config file
def scrape_homepage_main_menu_items(target_sys, headless_mode, username, password, user_roles, copied_files):
    try: 
        # Step 3: Scrape main menu items for each user role
        for user_role, copied_file in zip(user_roles, copied_files):

            driver_response = fdi.initialize_driver(headless_mode)

            menu_items = scrape_main_menu_items(target_sys, headless_mode, username, password, driver_response, user_role)
            

            # Step 4: Append menu items to each Excel file
            try:
                menu_items_array = menu_items
                # Step 3: Open the copied Excel file
                workbook = openpyxl.load_workbook(copied_file)
                sheet_name = 'Main_Menu_Items'  # Update with your sheet name if different
                sheet = workbook[sheet_name]                

                # Step 4: Append values to each column
                # Append each menu item in a new cell in the same column starting from the second row

                start_row = 2
                # Append each menu item in a new cell in the same column
                for row_num, menu_item in enumerate(menu_items_array, start=start_row):
                    sheet.cell(row=row_num, column=1, value=menu_item)

                # Step 5: Save the updated Excel file
                workbook.save(copied_file)
                print("\n")
                print(f"Data appended and saved to {copied_file}")
                # Print the user roles for verification
                print("User Roles:", user_role)                

                print("Menu items:", menu_items)
            except Exception as e:
                    print("An error occurred while fetching result:", e)
    except Exception as e:
        # Handle any exceptions that occur during the process
        print("An error occurred:", e)


# Function to fetch values from a specified column in the 'Main_Menu_Items' sheet.
def fetch_main_menu_items_from_excel(file_path, sheet_name='Main_Menu_Items', column=1):
    """
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from.
    :param column: Column number to read the values from (1-based index).
    :return: List of values in the specified column.
    """
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # List to hold the values
    values = []
    
    # Iterate over the rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
        for cell in row:
            if cell.value is not None:  # Ignore empty cells
                values.append(cell.value)
    
    return values


# Function to write to the tile names to the correct columns in the 'Menu Tiles Sheet'
def menu_tiles_write_to_specific_column(file_path, tile_items, menu_item, user_type, column_no) :
    # Step 3: Open the copied Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = 'Menu_Tiles'  # Update with your sheet name if different
    sheet = workbook[sheet_name]                

    # Step 4: Append values to each column
    # Append each menu item in a new cell in the same column starting from the second row

    start_row = 2
    # Append each menu item in a new cell in the same column
    for row_num, tile_item in enumerate(tile_items, start=start_row):
        sheet.cell(row=row_num, column=column_no, value=tile_item)

    

    # Step 5: Adjust the columns and save the updated Excel file
    adjust_column_widths(sheet)
    workbook.save(file_path)

    print("\n")
    print(f"Data appended and saved to {file_path}")
    # Print the user roles for verification
    print("User Roles:", user_type)                
    print("Menu Item: ", menu_item)
    print("Tile items: ", tile_items)
      

# Used to determine when all javascript fucntions have finished loading on in the DOM
def wait_for_page_load(driver, timeout=30):
    # Wait until document ready state is complete
    WebDriverWait(driver, timeout).until(
        lambda driver: driver.execute_script("return document.readyState") == "complete"
    )

# Function to scrape main menu menu tiles once loaded onto the webpage based on menu items that have been found for a given usertype/role
def scrape_main_menu_tiles(target_sys, headless_mode, username, password, driver_response, user_type, menu_items, file_path):
    def scrape_tiles(driver_response, menu_function_backup):
        max_attempts = 3
        attempts = 0
     
        while attempts < max_attempts:
            if attempts != 0: 
                menu_function_backup(target_sys, headless_mode, user_type, username, password, driver_response, True)
            try:

                # Use custom wait function to wait until page is fully loaded
                wait_for_page_load(driver_response)

                # Use explicit wait to ensure the 'uhTileFooterText' elements are present
                WebDriverWait(driver_response, 30).until(
                    EC.presence_of_all_elements_located((By.CLASS_NAME, 'uhTileFooterText'))
                )
                
                # Optionally, wait for JavaScript to finish executing
                time.sleep(10)  # Adjust if needed

                # Get the page source and parse it with BeautifulSoup
                page_source = driver_response.page_source
                soup = BeautifulSoup(page_source, 'html.parser')

                # Find all elements with the class 'uhTileFooterText'
                link_elements = soup.find_all(class_='uhTileFooterText')
                
                # Extract the text content of each element
                tile_names_set = {link.get_text(strip=True) for link in link_elements}

                # Convert the set to a list
                menu_tile_names_array = list(tile_names_set)

                 # Raise TimeoutException if the array is empty
                if not menu_tile_names_array:
                    raise TimeoutException("No tiles found.")

                return menu_tile_names_array
            
            except TimeoutException:
                print(f"Attempt {attempts + 1}: Timeout waiting for elements with class 'uhTileFooterText'. Retrying...")
                attempts += 1
                # Optionally, you can trigger a retry action here if needed
                # menu_function_backup(target_sys, headless_mode, user_type, username, password, driver_response, optional_error_catch)
     
        print("Maximum retry attempts reached. Unable to fetch tiles.")
        return []
        

    # Mapping menu items to their corresponding functions and column indices
    menu_function_map = {
        'Management': (mi.homePageLoadManagmentMenuItem, 1),
        'Data Processing': (mi.homePageLoadDataProcessingMenuItem, 2),
        'Bank': (mi.homePageLoadBankMenuItem, 3),
        'Self-Service': (mi.homePageLoadSelfServiceMenuItem, 4),
        'Reports': (mi.homePageLoadReportsMenuItem, 5),
        'Reports Dashboard': (mi.homePageLoadReportsDashboardMenuItem, 6),
        'About': (mi.homePageLoadAboutMenuItem, 7)
    }

    gs.loginAndLoadHomePage(target_sys, headless_mode, username, password, driver_response, True)
    gs.checkUserRoleAndSwitch(target_sys, headless_mode, user_type, username, password, driver_response, True)

    optional_error_catch = True

    # Iterate through the menu items and call the corresponding functions
    for menu_item in menu_items:
        #print(menu_item)
        # Get the function and column index from the map
        menu_function, column_index = menu_function_map.get(menu_item, (None, None))

        if menu_function is not None:
            menu_function(target_sys, headless_mode, user_type, username, password, driver_response, optional_error_catch)
            
            try:
                tile_items = scrape_tiles(driver_response, menu_function)
                menu_tiles_write_to_specific_column(file_path, tile_items, menu_item, user_type, column_index)
            except Exception as e:
                print("An error occurred while fetching result:", e)

    driver_response.quit()


# Function to call the 'scrape_main_menu_tiles' function on based on the usertype read from the config file aswell as
# the managment tiles written to the excel file via the  'scrape_homepage_main_menu_items' fucntion to fetch the homepage menu items. 
def scrape_homepage_main_menu_tiles(target_sys, headless_mode, username, password, user_roles, system_type): 
    #try: 
        # Ensures we write to the correct file when scraping the homepage_menu_tiles for source and target
        if system_type == True: 
            root_folder = 'source_system'
        else: 
            root_folder = 'target_system'

        i = 1
        
        file_paths = []
       
        for user_role in user_roles:
            
            file_path = f'..\\excel_sheets\\nav_homepage_items\\{root_folder}\\homepage_items_userrole_{i}.xlsx'  # Update with your file path

            # Here we append the file_paths so we can use use them at a later date when checking the difference between source and target. 
            file_paths.append(file_path)

            menu_items = fetch_main_menu_items_from_excel(file_path)
            print("Fetched Menu Items:", menu_items)
            
            driver_response = fdi.initialize_driver(headless_mode)

            scrape_main_menu_tiles(target_sys, headless_mode, username, password, driver_response, user_role, menu_items, file_path)

            i += 1

            #print(i)

        return file_paths
    #except Exception as e:
    #    # Handle any exceptions that occur during the process
    #    print("An error occurred:", e)
    #    driver_response.quit()


# Function to fetch values from a specified column in the 'Menu Tiles' sheet.
def fetch_main_menu_tiles_from_excel(file_path, column_no, sheet_name='Menu_Tiles'):
    """
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from.
    :param column: Column number to read the values from (1-based index).
    :return: List of values in the specified column.
    """
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # List to hold the values
    values = []
    
    # Iterate over the rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_no, max_col=column_no):
        for cell in row:
            if cell.value is not None:  # Ignore empty cells
                values.append(cell.value)
    
    return values


# Function to open up the tile based on the tilename passed it
def click_tiles_by_name(driver, tile_name):
    # Constructing the XPath to find the tile by its footer text content
    xpath = f"//div[@class='uhTileFooterText']//div[contains(text(), '{tile_name}')]/ancestor::div[@class='uhTileWrapper']"
    
    # Wait for the element to be clickable
    tile_elements = driver.find_elements(By.XPATH, xpath)
    
    if tile_elements:
        for tile_element in tile_elements:
            tile_element.click()
            time.sleep(3)  # Adjust as necessary to wait for page content to load after click


# Function to click on the close widget - acounts for case where the tile opens across the whole screen removing the rest of tiles from the DOM
def click_close_icon(driver):
    try:
        # Construct XPath to find the close icon
        xpath = "//div[@id='closeBtn' and contains(@class, 'closeBtn')]"
        
        # Wait until the close icon is present and clickable
        close_icon = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )

        # Scroll the element into view
        driver.execute_script("arguments[0].scrollIntoView(true);", close_icon)

        # Click the close icon
        close_icon.click()

        # Wait for a few seconds to ensure the close action completes
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element_located((By.XPATH, xpath))
        )
        print("Close icon clicked and popup closed.")
    
    except TimeoutException:
        print("Timeout: Close icon was not found or not clickable.")
    except NoSuchElementException:
        print("No such element: Close icon was not found on the page.")
    except ElementClickInterceptedException:
        print("Element click intercepted: The close icon could not be clicked.")
    except Exception as e:
        print(f"An error occurred: {e}")


# Function that will write to specific column based on the menu_item, tile_item and tab item passed to it in the sheet 'Menu_Tabs'
def menu_tabs_write_to_specific_column(file_path, row_num, menu_item, tile_item, tab_items, user_type, column_no_menu, column_no_tab):
    # Step 3: Open the copied Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = 'Menu_Tabs'  # Update with your sheet name if different
    sheet = workbook[sheet_name]

    # Step 4: Append values to each column starting from the second row
    start_row = 2

    # Iterate through menu items and tab items

   # Check if tab_items is a list of strings
    if isinstance(tab_items, list) and all(isinstance(i, str) for i in tab_items):
        # Concatenate tab items into a comma-separated string
        tab_items_str = ', '.join(tab_items)
        
        # Write the menu item to the specified column
        sheet.cell(row=row_num, column=column_no_menu, value=tile_item)
        
        # Write the tab items to the specified column
        sheet.cell(row=row_num, column=column_no_tab, value=tab_items_str)
        
        # Move to the next row
        row_num += 1
    else:
        raise ValueError("tab_items must be a list of strings")

    # Step 5: Adjust the columns and save the updated Excel file
    adjust_column_widths(sheet)
    workbook.save(file_path)

    print("\n")
    print(f"Data appended and saved to {file_path}")
    # Print the user roles for verification
    print("User Roles:", user_type)
    print("Menu Item: ", menu_item)
    print("Tile items: ", tile_item)
    print("Tab items: ", tab_items)


# Function used to fetch the names of the tabs present in the tile
def scrape_main_menu_tabs(target_sys, headless_mode, username, password, driver_response, user_type, menu_items, file_path):
    def scrape_tabs(driver_response):
        time.sleep(20)

        # Get the page source and parse it with BeautifulSoup
        page_source = driver_response.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find all elements with the class 'uhTileFooterText'
        link_elements = soup.find_all(class_='toptab tabshow')
        
        # Extract the text content of each element
        tab_names_set = {link.get_text(strip=True) for link in link_elements}

        # Convert the set to a list
        menu_tabs_names_array = list(tab_names_set)

        #print(tile_names_set)
        
        return menu_tabs_names_array

    # Mapping menu items to their corresponding functions
    menu_function_map = {
        'Management': (mi.homePageLoadManagmentMenuItem, 1),
        'Data Processing': (mi.homePageLoadDataProcessingMenuItem, 2),
        'Bank': (mi.homePageLoadBankMenuItem, 3),
        'Self Service': (mi.homePageLoadSelfServiceMenuItem, 4),
        'Reports': (mi.homePageLoadReportsMenuItem, 5),
        'Reports Dashboard': (mi.homePageLoadReportsDashboardMenuItem, 6),
        'About': (mi.homePageLoadAboutMenuItem, 7)
    }

    tab_column_map = {
        'Management': 2,
        'Data Processing': 4,
        'Bank': 6,
        'Self Service': 8,
        'Reports': 10,
        'Reports Dashboard': 12,
        'About': 14 
    }
    
    gs.loginAndLoadHomePage(target_sys, headless_mode, username, password, driver_response, True)
    gs.checkUserRoleAndSwitch(target_sys, headless_mode, user_type, username, password, driver_response, True)

    optional_error_catch = True 

    # Iterate through the menu items and call the corresponding functions
    for menu_item in menu_items:
        #print(menu_item)

        row_num = 2 # We define the row number here so that each iteration of the menu tiles is appended to a new row when calling the function to write to the 
                    # 'Menu tabs' sheet.
        
        
        # Get the function and column index from the map
        menu_function, column_index = menu_function_map.get(menu_item, (None, None))

        # Get the Column index for the 'Tab' columns in 'Menu_Tabs'
        column_index_tabs = tab_column_map.get(menu_item, (None, None))

        if menu_function is not None:
            menu_function(target_sys, headless_mode, user_type, username, password, driver_response, optional_error_catch)

            menu_tiles = fetch_main_menu_tiles_from_excel(file_path, column_index)

            for menu_tile in menu_tiles: 
                try: 
                    #print('menu_item', menu_item, 'menu_tile', menu_tile, 'tab_items', tab_items)
                    #print(menu_tiles)
                    #print(menu_tile)
                    click_tiles_by_name(driver_response, menu_tile)

                    tab_items = scrape_tabs(driver_response)
                    
                    click_close_icon(driver_response)
                                                     
                    menu_tabs_write_to_specific_column(file_path, row_num, menu_item, menu_tile, tab_items, user_type, column_index, column_index_tabs)

                    row_num += 1
                except Exception as e:
                    print("An error occurred while fetching result:", e)

    driver_response.quit()


# Function to recursivley call scrape_main_menu_tabs based on the user_role and menu items fecthed from the 'Menu Tiles' Sheet
def scrape_homepage_main_menu_tabs(target_sys, headless_mode, username, password, user_roles): 
    try: 
        i = 1
        # Step 3: Scrape main menu items for each user role   
        for user_role in user_roles:
            
            file_path = f'..\\excel_sheets\\nav_homepage_items\\source_system\\homepage_items_userrole_{i}.xlsx'  # Update with your file path
            
            #print(i)

            menu_items = fetch_main_menu_items_from_excel(file_path)
            print("Fetched Menu Items:", menu_items)
            
            driver_response = fdi.initialize_driver(headless_mode)

            scrape_main_menu_tabs(target_sys, headless_mode, username, password, driver_response, user_role, menu_items, file_path)
            
            i += 1
    except Exception as e:
        # Handle any exceptions that occur during the process
        print("An error occurred:", e)


# This function will compare the differences in one excel sheet against the other and create a new file with the menu items, 
# tiles and tabs that do not exist in both the source and target system, otherwise, it will not create a file and tell us that there are no differences.

# Sorting Values: Instead of directly comparing row by row, we'll sort the values within each column before comparing them. 
# This allows us to identify differences in the order of values without marking them as completely different.

# Using Sets for Comparison: Convert each column's values into sets after sorting. 
# Then, compare these sets between the source and target files. 
# Differences between sets will indicate values present in one file but not the other, or differences in the set of values.
def compare_excel_sheets(file1, file2, output_file):
    try:
        # Load the Excel files
        xls1 = pd.ExcelFile(file1)
        xls2 = pd.ExcelFile(file2)

        # Dictionary to store differences
        differences_dict = {}

        sheets = []
        output_results = False

        # Iterate through each sheet
        for sheet_name in xls1.sheet_names:
            df1 = pd.read_excel(xls1, sheet_name=sheet_name)
            df2 = pd.read_excel(xls2, sheet_name=sheet_name)

            # Normalize NaN values and strip whitespaces
            df1 = df1.fillna('').applymap(str.strip)
            df2 = df2.fillna('').applymap(str.strip)

            # Ensure columns are in a consistent order
            df1 = df1.reindex(sorted(df1.columns), axis=1)
            df2 = df2.reindex(sorted(df2.columns), axis=1)

            # Initialize DataFrame to store differences for current sheet
            diff_data = []

            # Iterate through each column
            for col in df1.columns:
                # Check if column exists in both DataFrames
                if col in df2.columns:
                    # Sort and convert values to sets for comparison
                    set1 = set(df1[col].dropna().astype(str).sort_values())
                    set2 = set(df2[col].dropna().astype(str).sort_values())

                    # Compare sets
                    if set1 != set2:
                        diff_data.append({
                            'Sheet': sheet_name,
                            'Column': col,
                            'Values in File 1': sorted(list(set1)),
                            'Values in File 2': sorted(list(set2))
                        })
                else:
                    # Handle case where column exists in df1 but not in df2
                    diff_data.append({
                        'Sheet': sheet_name,
                        'Column': col,
                        'Values in File 1': sorted(list(set1)),
                        'Values in File 2': 'Column does not exist in target file'
                    })

            # Check if differences were found
            if diff_data:
                differences_dict[sheet_name] = pd.DataFrame(diff_data)
                output_results = True
        
        if output_results:
            print(om.color_text("red") + "#####DIFFERENCES FOUND######" + om.color_text("white"))
            print(om.color_text("orange") + "Files: " + om.color_text("white"), file1)
            print("        and")
            print("       ", file2)
            print(om.color_text("blue") + "Sheets: " + om.color_text("white"), list(differences_dict.keys()))

        # Write differences to a new Excel file if differences were found
        if differences_dict:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name, diff_df in differences_dict.items():
                    diff_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    adjust_column_widths(worksheet)
            
            print(f"\nDifferences saved to: {output_file}")
        else:
           print(om.color_text("green") + "#####NO DIFFERENCES######" + om.color_text("white"))
           print(om.color_text("orange") + "Files: " + om.color_text("white"), file1)
           print("        and")
           print("       ", file2)

    except Exception as e:
        print(f"An error occurred: {e}")


# Function to get all main menu items and menu tiles for all user_roles passed into the config file for the system we want to fetch 
# the menu item objects from.

# To do's - Function to be able to fetch all tabs from their associated tiles under a given userrole 
def get_homepage_menu_objects_state(target_system, headless_mode, username, password, system_upgraded, config_file):
    # Fetches userroles and places into a list of strings
    user_roles = read_user_roles(config_file)
    
    if system_upgraded == False: 
        # This will create all files within the /source_system.
        copied_source_files = create_source_files(user_roles)

        scrape_homepage_main_menu_items(target_system, headless_mode, username, password, user_roles, copied_source_files)

        scrape_homepage_main_menu_tiles(target_system, headless_mode, username, password, user_roles, True)
    else:
        # This will create all files within the /target_system.
        copied_target_files = create_target_files(user_roles)

        scrape_homepage_main_menu_items(target_system, headless_mode, username, password, user_roles, copied_target_files)

        scrape_homepage_main_menu_tiles(target_system, headless_mode, username, password, user_roles, False)


# Function will compare the difference between the source files and target files generated for each user roles and produce a file
# outlining any differences in homepage objects found if any.
def system_compare_versions(): 
    # Specify the directory path to our source and target folders
    source_directory_path = '..\\excel_sheets\\nav_homepage_items\\source_system\\'  
    target_directory_path = '..\\excel_sheets\\nav_homepage_items\\target_system\\' 

    # Get the list of the source and target files
    source_files = list_files_in_directory(source_directory_path)
    target_files = list_files_in_directory(target_directory_path)

    i = 1 # used to pass the user_role number to the name of the file

    for source_file, target_file in zip(source_files, target_files):
        #print("Source file:", source_file)
        #print("Target file:", target_file)
        #print("\n")

        output_file = f'..\\excel_sheets\\nav_homepage_items\\differences\\user_role_{i}_differences.xlsx'

        # Save the differences to a new Excel file
        compare_excel_sheets(source_file, target_file, output_file)

        i += 1


# Test Environment
if __name__ == '__main__':

    # systems must have url's configured in /json_config/system_urls.json
    target_system = ''
    # If false (system state before being upgraded) - will create excel files for each userrole within the config with all objects found written to the associated files 
    # in '/source_system' 
    # If True (system state after being upgraded), will do the same, but will create andn write to files in '/target_system'
    #system_upgraded = True 
    headless_mode = False
    username = ''
    password = ''
    config_file = '..\\excel_sheets\\nav_homepage_items\\configs\\nav_homepage_items_config.xlsx'


    #print("#####Get Homepage menu objects#####")
    #print("If you would like to run for the old version of the system, enter 'o' otherwise, enter 'u' for upgraded version")
    #print("Otherwise enter 'd' to compare differences between the old and upgraded system: ")
    #user_input = input()
    #if user_input == 'o'.lower(): 
    #    system_upgraded = False 
    #    # Get Homepage Menu object - system before upgrade
    #    get_homepage_menu_objects_state(target_system, headless_mode, username, password, system_upgraded, config_file)
    #elif user_input == 'u'.lower(): 
    #    system_upgraded = True
    #    # Get Homepage Menu object - system after
    #    get_homepage_menu_objects_state(target_system, headless_mode, username, password, system_upgraded, config_file)
    #elif user_input == 'd'.lower(): 
    #    system_compare_versions()
    #else: 
    #    pass


    # Compares homepage objects from source version and upgrades version of system 
    #system_compare_versions()

    # Example usage
    #file1 = '..\\excel_sheets\\nav_homepage_items\\source_system\\homepage_items_userrole_1.xlsx'
    #file2 = '..\\excel_sheets\\nav_homepage_items\\target_system\\homepage_items_userrole_1.xlsx'
    #output_file = '..\\excel_sheets\\nav_homepage_items\\differences\\differences.xlsx'

    #print("\n")
    ## Save the differences to a new Excel file
    #compare_excel_sheets(file1, file2, output_file)

    # Fetches userroles and places into a list of strings
    user_roles = read_user_roles(config_file)

    scrape_homepage_main_menu_tabs(target_system, headless_mode, username, password, user_roles)